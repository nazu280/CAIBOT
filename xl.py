from openpyxl import load_workbook, Workbook

c_name = 1
c_id = 2
c_bot_id = 3
c_his_id = 4
c_inter_id = 5
c_botname = 6

wb = load_workbook("userDB.xlsx")
ws = wb.active

def loadFile():
    wb = load_workbook("userDB.xlsx")
    ws = wb.active
def saveFile():
    wb.save("userDB.xlsx")
    wb.close()

# 체크
def checkUserNum():
    print("xl.py - checkUserNum")
    loadFile()

    count = 0

    for row in range(2, ws.max_row+1):
        if ws.cell(row,c_name).value != None:
            count = count+1
        else:
            count = count
    return count

def checkFirstRow(_r):
    print("xl.py - checkFirstRow")
    loadFile()

    print("첫번째 빈 곳을 탐색")

    for row in range(2, ws.max_row + 1):
        if ws.cell(row,_r).value is None:
            return row
            break

    _result = ws.max_row+1

    saveFile()

    return _result

def checkUser(_name, _id):
    print("xl.py - checkUser")
    print(str(_name) + "<" + str(_id) + ">의 존재 여부 확인")
    print("")

    loadFile()

    userNum = checkUserNum()
    print("등록된 유저수: ", userNum)
    print("")

    print("고유번호 탐색")
    print("")

    for row in range(2, 3+userNum):

        print(row,"번째 줄 id: ", ws.cell(row,c_id).value)
        print("입력된 id: ", hex(_id))
        print("고유번호정보와 일치 여부: ", ws.cell(row, c_id).value == hex(_id))
        print("")

        if ws.cell(row,c_id).value == hex(_id):
            print("등록된 고유번호를 발견")
            print("등록된 값의 위치: ",  row, "번째 줄")
            print("")

            saveFile()

            return True, row
            break
        else:
            print("등록된 정보를 탐색 실패, 재탐색 실시")

    saveFile()
    print("발견 실패")

    return False, None

# 셋업
def set1(_name, _id):
    print("xl.py - set")

    loadFile()
    global userExistance
    global _row1
    userExistance, _row1 = checkUser(_name, _id)
    if userExistance:
        print("기존 유저 데이터 수정 시작")

        ws.cell(row=_row1, column=c_name, value=_name)
        print("이름 추가 | ",  ws.cell(_row1,c_name).value)
        ws.cell(row=_row1, column=c_id, value =hex(_id))
        print("고유번호 추가 | ", ws.cell(_row1,c_id).value)

        print("")

        saveFile()

        print("데이터 수정 완료")
    else:
        _row = checkFirstRow(1)
        print("첫번째 빈곳: ", _row)
        print("")

        print("신규 유저 데이터 추가 시작")

        ws.cell(row=_row, column=c_name, value=_name)
        print("이름 추가 | ",  ws.cell(_row,c_name).value)
        ws.cell(row=_row, column=c_id, value =hex(_id))
        print("고유번호 추가 | ", ws.cell(_row,c_id).value)

        print("")

        saveFile()

        print("데이터 추가 완료")

def set2(_bot_id, _his_id):
    print("xl.py - set")

    loadFile()
    if userExistance:
        print("기존 유저 데이터 수정 시작")

        ws.cell(row=_row1, column=c_bot_id, value = _bot_id)
        print("봇 아이디 추가 | ", ws.cell(_row1,c_bot_id).value)
        ws.cell(row=_row1, column=c_his_id, value = _his_id)
        print("히스토리 아이디 추가 | ", ws.cell(_row1,c_his_id).value)
        print("")

        saveFile()

        print("데이터 수정 완료")
    else:
        _row = checkFirstRow(3)
        print("첫번째 빈곳: ", _row)
        print("")

        print("신규 유저 데이터 추가 시작")

        ws.cell(row=_row, column=c_bot_id, value = _bot_id)
        print("봇 아이디 추가 | ", ws.cell(_row,c_bot_id).value)
        ws.cell(row=_row, column=c_his_id, value = _his_id)
        print("히스토리 아이디 추가 | ", ws.cell(_row,c_his_id).value)

        print("")

        saveFile()

        print("데이터 추가 완료")

def set3(_inter_id, _botname):
    print("xl.py - set")

    loadFile()
    if userExistance:
        print("기존 유저 데이터 수정 시작")

        ws.cell(row=_row1, column=c_inter_id, value = _inter_id)
        print("내부 아이디 추가 | ", ws.cell(_row1,c_inter_id).value)
        ws.cell(row=_row1, column=c_botname, value = _botname)
        print("봇이름 추가 | ", ws.cell(_row1,c_botname).value)

        print("")

        saveFile()

        print("데이터 수정 완료")
    else:
        _row = checkFirstRow(5)
        print("첫번째 빈곳: ", _row)
        print("")

        print("신규 유저 데이터 추가 시작")

        ws.cell(row=_row, column=c_inter_id, value = _inter_id)
        print("내부 아이디 추가 | ", ws.cell(_row,c_inter_id).value)
        ws.cell(row=_row, column=c_botname, value = _botname)
        print("봇이름 추가 | ", ws.cell(_row1,c_botname).value)

        print("")

        saveFile()

        print("데이터 추가 완료")

# 유저 정보
def userInfo(_row):
    loadFile()

    _bot_id = ws.cell(_row,c_bot_id).value
    _his_id = ws.cell(_row,c_his_id).value
    _inter_id = ws.cell(_row,c_inter_id).value
    _bot_name = ws.cell(_row,c_botname).value

    print("봇 아이디: ", _bot_id)
    print("대화 아이디: ", _his_id)
    print("내부 아이디: ", _inter_id)
    print("봇 이름: ", _bot_name)

    saveFile()

    return _bot_id, _his_id, _inter_id, _bot_name